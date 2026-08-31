import io
import json
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from typing import Dict, Any, List

from reportlab.pdfbase.pdfmetrics import registerFontFamily

# Register Unicode Font for full Latin + Devanagari (Hindi / Marathi) rendering
_DOC_FONT = "Helvetica"
_DOC_FONT_BOLD = "Helvetica-Bold"

_FONT_DIR = os.path.join(os.path.dirname(__file__), "..", "fonts")
_FONT_CANDIDATES = [
    os.path.join(_FONT_DIR, "Nirmala.ttf"),
    os.path.join(_FONT_DIR, "NotoSansDevanagari-Regular.ttf"),
    "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
    "C:/Windows/Fonts/Nirmala.ttf"
]

for _fc in _FONT_CANDIDATES:
    if os.path.exists(_fc):
        try:
            pdfmetrics.registerFont(TTFont("DocRefineFont", _fc))
            registerFontFamily(
                "DocRefineFont",
                normal="DocRefineFont",
                bold="DocRefineFont",
                italic="DocRefineFont",
                boldItalic="DocRefineFont"
            )
            _DOC_FONT = "DocRefineFont"
            _DOC_FONT_BOLD = "DocRefineFont"
            print(f"[Exporters] Registered Unicode font from {_fc}")
            break
        except Exception as _f_err:
            print(f"[Exporters] Could not register {_fc}: {_f_err}")


def export_json(case_data: Dict[str, Any], documents: List[Dict[str, Any]], extractions: List[Dict[str, Any]]) -> bytes:
    """Generates canonical machine-readable JSON export."""
    payload = {
        "export_version": "2.0.0",
        "case": case_data,
        "document_count": len(documents),
        "documents": []
    }
    for doc in documents:
        doc_id = doc.get("id")
        ext = next((e for e in extractions if e.get("document_id") == doc_id), {})
        payload["documents"].append({
            "document_id": doc_id,
            "filename": doc.get("filename_safe"),
            "status": doc.get("status"),
            "quality_score": doc.get("quality_score"),
            "quality_band": doc.get("quality_band"),
            "quality_flags": doc.get("quality_flags_json"),
            "schema_name": ext.get("schema_name"),
            "overall_confidence": ext.get("overall_confidence"),
            "review_status": ext.get("status"),
            "approved_at": ext.get("approved_at"),
            "approved_by": ext.get("approved_by"),
            "fields": ext.get("fields", [])
        })
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


def export_csv(case_data: Dict[str, Any], documents: List[Dict[str, Any]], extractions: List[Dict[str, Any]]) -> bytes:
    """
    Generates a clean, well-structured business CSV export with UTF-8 BOM.
    Readable column headers and properly formatted extracted values.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    case_name = case_data.get("name", "Review Case")

    # Clean Business Headers
    writer.writerow([
        "Case Name",
        "Document Name",
        "Document Type",
        "Field Name",
        "Extracted Value",
        "Raw OCR Snippet",
        "Confidence Score",
        "Confidence State",
        "Validation Status",
        "Review Status",
        "Document Quality Score"
    ])

    for doc in documents:
        doc_id = doc.get("id")
        fname = doc.get("filename_safe", "Document")
        q_score = doc.get("quality_score", "N/A")
        ext = next((e for e in extractions if e.get("document_id") == doc_id), {})
        doc_type = (ext.get("schema_name") or "generic").upper()
        rev_status = ext.get("status", "needs_review").upper()
        fields = ext.get("fields", [])

        if not fields:
            writer.writerow([
                case_name,
                fname,
                doc_type,
                "[NO_FIELDS_EXTRACTED]",
                "",
                "",
                "0%",
                "Red",
                "missing",
                rev_status,
                f"{q_score}/100"
            ])
        else:
            for f in fields:
                val = f.get("normalized_value") or f.get("raw_value") or ""
                raw_v = f.get("raw_value") or ""
                score = f"{f.get('field_score', 0)}%"
                state = f.get("confidence_state", "Yellow")
                val_status = f.get("validation_status", "valid")
                f_rev = (f.get("review_state") or rev_status).upper()

                writer.writerow([
                    case_name,
                    fname,
                    doc_type,
                    f.get("label") or f.get("field_key", ""),
                    val,
                    raw_v,
                    score,
                    state,
                    val_status,
                    f_rev,
                    f"{q_score}/100"
                ])

    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")


def export_xlsx(case_data: Dict[str, Any], documents: List[Dict[str, Any]], extractions: List[Dict[str, Any]]) -> bytes:
    """
    Generates an executive, beautifully formatted multi-sheet Excel workbook.
    Sheet 1: Executive Summary & All Extracted Fields visible immediately.
    Sheet 2: Detailed OCR Forensic Audit & Reason Codes.
    """
    wb = openpyxl.Workbook()

    # Style definitions
    title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

    section_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    meta_label_font = Font(name="Segoe UI", size=10, bold=True, color="475569")
    meta_val_font = Font(name="Segoe UI", size=10, color="0F172A")
    meta_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

    data_font = Font(name="Segoe UI", size=10, color="1E293B")
    data_font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Status & Confidence Fills
    green_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    green_font = Font(name="Segoe UI", size=10, bold=True, color="065F46")

    amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    amber_font = Font(name="Segoe UI", size=10, bold=True, color="92400E")

    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # ─────────────────────────────────────────────────────────────
    # SHEET 1: Executive Summary & Extracted Fields
    # ─────────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Executive Summary & Fields"
    ws_summary.views.sheetView[0].showGridLines = True

    # 1. Top Title Banner
    ws_summary.merge_cells("A1:G1")
    title_cell = ws_summary.cell(row=1, column=1, value="DocRefine — Verified Document Extraction Audit Report")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 36

    # 2. Metadata Block
    meta_info = [
        ("Case Name", case_data.get("name", "N/A"), "Export Timestamp", case_data.get("updated_at", "N/A")),
        ("Case ID", case_data.get("id", "N/A"), "Total Documents", len(documents)),
        ("Review Status", "Verified & Human-Approved" if any(e.get("status") == "approved" for e in extractions) else "Pending Review", "Platform Version", "DocRefine AI v2.8")
    ]

    for r_offset, (k1, v1, k2, v2) in enumerate(meta_info):
        r_idx = 3 + r_offset
        ws_summary.row_dimensions[r_idx].height = 20

        # Col A-B
        c_k1 = ws_summary.cell(row=r_idx, column=1, value=k1)
        c_k1.font = meta_label_font
        c_k1.fill = meta_fill
        c_k1.border = thin_border

        c_v1 = ws_summary.cell(row=r_idx, column=2, value=str(v1))
        c_v1.font = meta_val_font
        c_v1.fill = white_fill
        c_v1.border = thin_border

        # Col C-D
        c_k2 = ws_summary.cell(row=r_idx, column=4, value=k2)
        c_k2.font = meta_label_font
        c_k2.fill = meta_fill
        c_k2.border = thin_border

        c_v2 = ws_summary.cell(row=r_idx, column=5, value=str(v2))
        c_v2.font = meta_val_font
        c_v2.fill = white_fill
        c_v2.border = thin_border

    # 3. Documents Overview Table
    doc_headers = [
        "Filename", "Document Type", "Quality Score", "Quality Band",
        "Overall Confidence", "Extracted Fields", "Review Status"
    ]
    doc_header_row = 7
    ws_summary.row_dimensions[doc_header_row].height = 24

    for col_idx, text in enumerate(doc_headers, 1):
        cell = ws_summary.cell(row=doc_header_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    curr_row = doc_header_row + 1
    for doc in documents:
        doc_id = doc.get("id")
        ext = next((e for e in extractions if e.get("document_id") == doc_id), {})
        fields = ext.get("fields", [])
        field_count_str = f"{len(fields)} fields extracted" if fields else "None"
        q_score = doc.get("quality_score", 0)
        q_band = doc.get("quality_band", "acceptable")
        conf = f"{ext.get('overall_confidence', 0)}%"
        st = ext.get("status", "needs_review").upper()

        ws_summary.row_dimensions[curr_row].height = 22
        row_fill = alt_fill if curr_row % 2 == 0 else white_fill

        values = [
            (doc.get("filename_safe", "document"), Alignment(horizontal="left", vertical="center")),
            ((ext.get("schema_name") or "generic").upper(), Alignment(horizontal="center", vertical="center")),
            (f"{q_score}/100", Alignment(horizontal="center", vertical="center")),
            (q_band.title(), Alignment(horizontal="center", vertical="center")),
            (conf, Alignment(horizontal="center", vertical="center")),
            (field_count_str, Alignment(horizontal="center", vertical="center")),
            (st, Alignment(horizontal="center", vertical="center"))
        ]

        for col_idx, (val, align) in enumerate(values, 1):
            cell = ws_summary.cell(row=curr_row, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = align
            cell.border = thin_border

            if col_idx == 7:
                if st == "APPROVED":
                    cell.fill = green_fill
                    cell.font = green_font
                elif st in ("NEEDS_REVIEW", "PENDING"):
                    cell.fill = amber_fill
                    cell.font = amber_font
                else:
                    cell.fill = red_fill
                    cell.font = red_font
        curr_row += 1

    # 4. Prominent Extracted Data Section on Sheet 1
    curr_row += 2
    ws_summary.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
    section_cell = ws_summary.cell(row=curr_row, column=1, value="VERIFIED EXTRACTED FIELDS BY DOCUMENT")
    section_cell.font = section_font
    section_cell.fill = section_fill
    section_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_summary.row_dimensions[curr_row].height = 28
    curr_row += 1

    ext_field_headers = [
        "Document", "Document Type", "Field Name", "Extracted Value",
        "Confidence Score", "Validation Status", "Review Status"
    ]
    ws_summary.row_dimensions[curr_row].height = 24
    for col_idx, text in enumerate(ext_field_headers, 1):
        cell = ws_summary.cell(row=curr_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    curr_row += 1

    has_any_fields = False
    for doc in documents:
        doc_id = doc.get("id")
        fname = doc.get("filename_safe", "document")
        ext = next((e for e in extractions if e.get("document_id") == doc_id), {})
        doc_type = (ext.get("schema_name") or "generic").upper()
        doc_fields = ext.get("fields", [])

        for f in doc_fields:
            has_any_fields = True
            ws_summary.row_dimensions[curr_row].height = 20
            row_fill = alt_fill if curr_row % 2 == 0 else white_fill

            norm_val = str(f.get("normalized_value") or f.get("raw_value") or "[MISSING]")
            score_num = f.get("field_score", 0)
            score_str = f"{score_num}%"
            state = f.get("confidence_state", "Yellow")
            val_st = f.get("validation_status", "valid").upper()
            rev_st = (f.get("review_state") or ext.get("status", "needs_review")).upper()

            f_values = [
                (fname, Alignment(horizontal="left", vertical="center")),
                (doc_type, Alignment(horizontal="center", vertical="center")),
                (f.get("label") or f.get("field_key", ""), Alignment(horizontal="left", vertical="center")),
                (norm_val, Alignment(horizontal="left", vertical="center")),
                (score_str, Alignment(horizontal="center", vertical="center")),
                (val_st, Alignment(horizontal="center", vertical="center")),
                (rev_st, Alignment(horizontal="center", vertical="center"))
            ]

            for col_idx, (val, align) in enumerate(f_values, 1):
                cell = ws_summary.cell(row=curr_row, column=col_idx, value=val)
                cell.font = data_font_bold if col_idx == 4 else data_font
                cell.fill = row_fill
                cell.alignment = align
                cell.border = thin_border

                if col_idx == 5: # Score
                    if state == "Green":
                        cell.fill = green_fill
                        cell.font = green_font
                    elif state == "Yellow":
                        cell.fill = amber_fill
                        cell.font = amber_font
                    else:
                        cell.fill = red_fill
                        cell.font = red_font
            curr_row += 1

    if not has_any_fields:
        ws_summary.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
        c_none = ws_summary.cell(row=curr_row, column=1, value="No extracted fields available. Please run extraction on documents.")
        c_none.font = data_font
        c_none.alignment = Alignment(horizontal="center", vertical="center")
        curr_row += 1

    # ─────────────────────────────────────────────────────────────
    # SHEET 2: Forensic Field Audit & Raw OCR Snippets
    # ─────────────────────────────────────────────────────────────
    ws_fields = wb.create_sheet(title="Forensic Field Audit")
    ws_fields.views.sheetView[0].showGridLines = True

    field_headers = [
        "Document Filename", "Document Type", "Field Key", "Field Label", "Normalized Value",
        "Raw OCR Snippet", "Confidence Score", "Confidence State", "Validation Status", "Review State"
    ]
    ws_fields.row_dimensions[1].height = 26

    for col_idx, text in enumerate(field_headers, 1):
        cell = ws_fields.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    current_field_row = 2
    for doc in documents:
        doc_id = doc.get("id")
        fname = doc.get("filename_safe", "document")
        ext = next((e for e in extractions if e.get("document_id") == doc_id), {})
        schema_type = (ext.get("schema_name") or "generic").upper()
        doc_fields = ext.get("fields", [])

        for f in doc_fields:
            ws_fields.row_dimensions[current_field_row].height = 20
            row_fill = alt_fill if current_field_row % 2 == 0 else white_fill

            norm_val = str(f.get("normalized_value") or f.get("raw_value") or "[MISSING]")
            raw_val = str(f.get("raw_value") or "")
            score_num = f.get("field_score", 0)
            score_str = f"{score_num}%"
            state = f.get("confidence_state", "Red")
            val_st = f.get("validation_status", "valid")
            rev_st = f.get("review_state", "needs_review")

            field_vals = [
                (fname, Alignment(horizontal="left", vertical="center")),
                (schema_type, Alignment(horizontal="center", vertical="center")),
                (f.get("field_key", ""), Alignment(horizontal="left", vertical="center")),
                (f.get("label", ""), Alignment(horizontal="left", vertical="center")),
                (norm_val, Alignment(horizontal="left", vertical="center")),
                (raw_val, Alignment(horizontal="left", vertical="center")),
                (score_str, Alignment(horizontal="center", vertical="center")),
                (state, Alignment(horizontal="center", vertical="center")),
                (val_st, Alignment(horizontal="center", vertical="center")),
                (rev_st, Alignment(horizontal="center", vertical="center"))
            ]

            for col_idx, (val, align) in enumerate(field_vals, 1):
                cell = ws_fields.cell(row=current_field_row, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = row_fill
                cell.alignment = align
                cell.border = thin_border

                if col_idx == 8: # Confidence State
                    if state == "Green":
                        cell.fill = green_fill
                        cell.font = green_font
                    elif state == "Yellow":
                        cell.fill = amber_fill
                        cell.font = amber_font
                    else:
                        cell.fill = red_fill
                        cell.font = red_font

            current_field_row += 1

    # Auto-adjust column widths for both sheets
    for ws in [ws_summary, ws_fields]:
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                val = str(cell.value or "")
                if "\n" in val:
                    val = max(val.split("\n"), key=len)
                if len(val) > max_len and cell.row > 1: # Ignore title banner
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(min(max_len + 4, 45), 16)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _has_devanagari(text: str) -> bool:
    import re
    return bool(re.search(r'[\u0900-\u097F]', str(text or '')))


def export_pdf(case_data: Dict[str, Any], documents: List[Dict[str, Any]], extractions: List[Dict[str, Any]]) -> bytes:
    """
    Generates an executive, publication-grade PDF case audit report with
    Devanagari Unicode support, high-contrast metadata cards, and structured tables.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocRefineTitle',
        parent=styles['Heading1'],
        fontName=_DOC_FONT_BOLD,
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#FFFFFF')
    )

    subtitle_style = ParagraphStyle(
        'DocRefineSubtitle',
        parent=styles['Normal'],
        fontName=_DOC_FONT,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#94A3B8')
    )

    h2_style = ParagraphStyle(
        'DocRefineH2',
        parent=styles['Heading2'],
        fontName=_DOC_FONT_BOLD,
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#0F172A'),
        spaceBefore=8,
        spaceAfter=4
    )

    meta_label = ParagraphStyle(
        'DocRefineMetaLabel',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#475569')
    )

    meta_val = ParagraphStyle(
        'DocRefineMetaVal',
        parent=styles['Normal'],
        fontName=_DOC_FONT,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#0F172A')
    )

    th_style = ParagraphStyle(
        'DocRefineTH',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#FFFFFF'),
        alignment=1  # Center
    )

    th_style_left = ParagraphStyle(
        'DocRefineTHLeft',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#FFFFFF')
    )

    cell_style = ParagraphStyle(
        'DocRefineCell',
        parent=styles['Normal'],
        fontName=_DOC_FONT,
        fontSize=8,
        leading=10.5,
        textColor=colors.HexColor('#1E293B')
    )

    cell_bold = ParagraphStyle(
        'DocRefineCellBold',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8,
        leading=10.5,
        textColor=colors.HexColor('#0F172A')
    )

    cell_score_green = ParagraphStyle(
        'DocRefineScoreGreen',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8,
        leading=10.5,
        textColor=colors.HexColor('#065F46'),
        alignment=1
    )

    cell_score_amber = ParagraphStyle(
        'DocRefineScoreAmber',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8,
        leading=10.5,
        textColor=colors.HexColor('#92400E'),
        alignment=1
    )

    cell_score_red = ParagraphStyle(
        'DocRefineScoreRed',
        parent=styles['Normal'],
        fontName=_DOC_FONT_BOLD,
        fontSize=8,
        leading=10.5,
        textColor=colors.HexColor('#991B1B'),
        alignment=1
    )

    def _make_para(text: str, default_st: ParagraphStyle) -> Paragraph:
        return Paragraph(str(text or ''), default_st)

    story = []

    # 1. Header Banner
    banner_content = [
        [Paragraph("DocRefine — Verified Document Rescue & Extraction Report", title_style)],
        [Paragraph("AI-Powered Multilingual Extraction • OCR Rescue • Maker-Checker Verification", subtitle_style)]
    ]
    banner_table = Table(banner_content, colWidths=[540])
    banner_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#0F172A')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 8))

    # 2. Case Metadata Card (2x2 table)
    is_approved = any(e.get("status") == "approved" for e in extractions)
    status_text = "Verified & Human-Approved" if is_approved else "Pending Verification"

    meta_table_data = [
        [
            Paragraph("Case Name:", meta_label),
            _make_para(case_data.get("name", "Review Case"), meta_val),
            Paragraph("Case ID:", meta_label),
            Paragraph(str(case_data.get("id", "N/A")), meta_val)
        ],
        [
            Paragraph("Review Status:", meta_label),
            Paragraph(status_text, meta_val),
            Paragraph("Total Documents:", meta_label),
            Paragraph(f"{len(documents)} file(s)", meta_val)
        ],
        [
            Paragraph("Export Time:", meta_label),
            Paragraph(str(case_data.get("updated_at", "UTC")), meta_val),
            Paragraph("Engine Version:", meta_label),
            Paragraph("DocRefine Verification Lab v2.8", meta_val)
        ]
    ]
    meta_table = Table(meta_table_data, colWidths=[90, 180, 90, 180])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 10))

    # 3. Documents & Extracted Fields
    for doc_idx, doc_item in enumerate(documents, 1):
        doc_id = doc_item.get("id")
        fname = doc_item.get("filename_safe", f"Document #{doc_idx}")
        ext = next((e for e in extractions if e.get("document_id") == doc_id), {})
        doc_type = (ext.get("schema_name") or "generic").upper().replace("_", " ")
        rev_status = (ext.get("status") or "needs_review").upper()
        q_score = doc_item.get("quality_score", 0)
        q_band = (doc_item.get("quality_band") or "acceptable").title()
        conf_overall = f"{ext.get('overall_confidence', 0)}%"

        # Document Header Card
        doc_summary_text = (
            f"<b>{fname}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Type: <b>{doc_type}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Quality: <b>{q_score}/100 ({q_band})</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Confidence: <b>{conf_overall}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Status: <b>{rev_status}</b>"
        )
        doc_hdr_table = Table([[Paragraph(doc_summary_text, ParagraphStyle('DocHdr', fontName='Helvetica', textColor=colors.HexColor('#FFFFFF'), fontSize=8.5, leading=11))]], colWidths=[540])
        doc_hdr_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1E293B')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(doc_hdr_table)

        # Fields Table
        fields = ext.get("fields", [])
        if fields:
            table_data = [[
                Paragraph("Field Name", th_style_left),
                Paragraph("Extracted Value", th_style_left),
                Paragraph("Confidence Score", th_style),
                Paragraph("Validation Status", th_style)
            ]]

            table_styles = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]

            for row_idx, f in enumerate(fields, 1):
                val = str(f.get("normalized_value") or f.get("raw_value") or "[MISSING]")
                state = f.get("confidence_state", "Yellow")
                score_num = f.get("field_score", 0)
                score_str = f"{score_num}% ({state})"
                val_status = (f.get("validation_status") or "valid").upper()

                if state == "Green":
                    score_p = Paragraph(score_str, cell_score_green)
                elif state == "Yellow":
                    score_p = Paragraph(score_str, cell_score_amber)
                else:
                    score_p = Paragraph(score_str, cell_score_red)

                # Alternating row background
                bg_color = colors.HexColor('#F8FAFC') if row_idx % 2 == 0 else colors.HexColor('#FFFFFF')
                table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))

                label_text = f.get("label") or f.get("field_key", "")
                table_data.append([
                    _make_para(label_text, cell_bold),
                    _make_para(val, cell_style),
                    score_p,
                    Paragraph(val_status, ParagraphStyle(f'ValSt_{row_idx}', fontName='Helvetica-Bold', fontSize=8, alignment=1, textColor=colors.HexColor('#059669') if val_status == 'VALID' else colors.HexColor('#DC2626')))
                ])

            t = Table(table_data, colWidths=[130, 230, 95, 85])
            t.setStyle(TableStyle(table_styles))
            story.append(t)
        else:
            no_fields_p = Paragraph("<i>No structured fields extracted for this document.</i>", ParagraphStyle('NoF', fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#64748B'), alignment=1))
            story.append(Table([[no_fields_p]], colWidths=[540], style=[
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))

        story.append(Spacer(1, 10))

    # 4. Footer Note
    footer_text = "DocRefine AI Intelligence • Confidential Verification Audit Report • Export generated via DocRefine API"
    story.append(Paragraph(footer_text, ParagraphStyle('DocFooter', fontName='Helvetica', fontSize=7.5, leading=9, textColor=colors.HexColor('#94A3B8'), alignment=1)))

    doc.build(story)
    return buffer.getvalue()
